from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from ivenn import IVenn, Set


def small_controller() -> IVenn:
    return IVenn(
        Set("A", ["a1", "a2", "ab1", "abc1"]),
        Set("B", ["b1", "ab1", "abc1"]),
        Set("C", ["c1", "abc1"]),
    )


def six_set_controller() -> IVenn:
    return IVenn(
        Set("A", ["a", "ab", "abcdef"]),
        Set("B", ["b", "ab", "abcdef"]),
        Set("C", ["c", "abcdef"]),
        Set("D", ["d", "abcdef"]),
        Set("E", ["e", "abcdef"]),
        Set("F", ["f", "abcdef"]),
    )


def test_init_rejects_zero_sets() -> None:
    with pytest.raises(ValueError, match="at least one Set"):
        IVenn()


def test_init_rejects_more_than_six_sets() -> None:
    sets = [Set(f"S{i}", [str(i)]) for i in range(7)]
    with pytest.raises(ValueError, match="supports up to 6 sets"):
        IVenn(*sets)


def test_from_excel(tmp_path: Path) -> None:
    file_path = tmp_path / "input.xlsx"
    pd.DataFrame(
        {
            "Col A": ["12", "12.0", " X ", None, "", "x"],
            "Col B": ["7.0", "7", "Y", "  y  ", None, ""],
        }
    ).to_excel(file_path, index=False)

    ivenn = IVenn.from_excel(str(file_path))

    assert ivenn.labels["A"] == "Col A"
    assert ivenn.labels["B"] == "Col B"
    assert ivenn.sets["A"] == {"12", "X", "x"}
    assert ivenn.sets["B"] == {"7", "Y", "y"}


def test_set_theme() -> None:
    ivenn = six_set_controller()

    ivenn.set_theme("Default")
    assert ivenn.theme == "Default"

    ivenn.set_theme("#111111", "#222222", "#333333", "#444444", "#555555", "#666666")
    assert ivenn.theme == "Custom"
    assert ivenn.custom_theme["A"] == "#111111"
    assert ivenn.custom_theme["F"] == "#666666"


def test_set_theme_rejects_invalid_input() -> None:
    ivenn = six_set_controller()

    with pytest.raises(ValueError, match="Unknown theme"):
        ivenn.set_theme("NotATheme")

    with pytest.raises(ValueError, match="Colours must be hex strings"):
        ivenn.set_theme("#111111", "#222222", "bad", "#444444", "#555555", "#666666")

    with pytest.raises(TypeError, match="set_theme expects"):
        ivenn.set_theme("#111111", "#222222")


def test_set_unions_list_and_navigation() -> None:
    ivenn = six_set_controller()
    ivenn.set_unions("ab; cde, f")

    assert ivenn.union_mode == "list"
    assert len(ivenn.union_states) == 3

    ivenn._next()
    assert ivenn._collapsed_sets()["AB"] == (ivenn.sets["A"] | ivenn.sets["B"])

    ivenn._next()
    assert ivenn._collapsed_sets()["CDE"] == (ivenn.sets["C"] | ivenn.sets["D"] | ivenn.sets["E"])

    ivenn._prev()
    assert ivenn.current_index == 1

    ivenn._stop()
    assert ivenn.current_index == 0
    assert ivenn.union_states == [[]]


def test_set_unions_list_rejects_invalid_input() -> None:
    ivenn = small_controller()

    with pytest.raises(ValueError, match="Unknown set"):
        ivenn.set_unions("az")

    with pytest.raises(ValueError, match="Overlapping unions"):
        ivenn.set_unions("ab,bc")


def test_set_unions_tree() -> None:
    ivenn = small_controller()
    ivenn.set_unions("((A,B),C)")

    assert ivenn.union_mode == "tree"
    assert ivenn.union_states == [[], [set("AB")], [set("ABC")]]


def test_set_unions_tree_rejects_invalid_input() -> None:
    ivenn = small_controller()

    with pytest.raises(ValueError, match="Invalid character"):
        ivenn.set_unions("((A,B),1)")

    with pytest.raises(ValueError, match="Unknown set"):
        ivenn.set_unions("((A,Z),B)")

    with pytest.raises(ValueError, match="Each tree union group must contain at least 2 items"):
        ivenn.set_unions("(A)")


def test_export_sets(tmp_path: Path) -> None:
    ivenn = small_controller()
    out_file = tmp_path / "sets.xlsx"

    ivenn.export_sets(str(out_file))

    workbook = openpyxl.load_workbook(out_file)
    worksheet = workbook.active
    headers = [worksheet.cell(row=1, column=i).value for i in range(1, worksheet.max_column + 1)]

    assert headers == ["A", "B", "C"]


def test_export_intersections(tmp_path: Path) -> None:
    ivenn = small_controller()
    out_file = tmp_path / "intersections.xlsx"

    ivenn.export_intersections(str(out_file), min_size=1)

    workbook = openpyxl.load_workbook(out_file)
    worksheet = workbook.active
    headers = [worksheet.cell(row=1, column=i).value for i in range(1, worksheet.max_column + 1)]

    assert "A" in headers
    assert "A ∩ B" in headers
    assert "A ∩ B ∩ C" in headers


def test_render_svg(tmp_path: Path) -> None:
    ivenn = small_controller()
    template_base = Path(__file__).resolve().parents[1] / "src" / "ivenn" /  "diagrams"
    out_file = tmp_path / "rendered.svg"

    ivenn.export_svg(str(out_file), base_dir=str(template_base))

    text = out_file.read_text(encoding="utf-8")
    assert "A" in text
    assert "B" in text
    assert "C" in text


def test_output_consistency(tmp_path: Path) -> None:
    ivenn = small_controller()
    out_file = tmp_path / "intersections.xlsx"

    ivenn.export_intersections(str(out_file), min_size=1)

    workbook = openpyxl.load_workbook(out_file)
    worksheet = workbook.active
    headers = [worksheet.cell(row=1, column=i).value for i in range(1, worksheet.max_column + 1)]
    ab_column = headers.index("A ∩ B") + 1

    exported = []
    row = 2
    while True:
        value = worksheet.cell(row=row, column=ab_column).value
        if value in (None, ""):
            break
        exported.append(str(value))
        row += 1

    assert sorted(ivenn.get_intersection("AB")) == ["ab1"]
    key = ivenn._normalise_intersection_lookup("AB")
    assert sorted(ivenn._region_elements()[key]) == ["ab1"]
    assert sorted(exported) == ["ab1"]