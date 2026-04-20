from contextlib import contextmanager
from importlib.resources import as_file, files
from itertools import combinations
import os
import re

import pandas as pd
from lxml import etree
from openpyxl import Workbook

from .models import Set
from .themes import SET_COLOUR_THEMES, get_theme, theme_names, validate_theme

SVG_NS = {"svg": "http://www.w3.org/2000/svg"}


class IVenn:
    """Create, analyse, and export InteractiVenn-style diagrams for up to 6 sets.

    The class stores up to six input ``Set``s, supports optional unions,
    computes intersection elements and counts, and exports the current diagram
    as either SVG, PNG, or Excel files.

    ## Example:


        v = IVenn(
            Set("A", [1, 2, 3]),
            Set("B", [3, 4, 5]),
        )
    """

    def __init__(self, *sets: Set):
        """Create an ``IVenn`` object from one to six ``Set`` objects.

        :param sets: Between 1 and 6 ``Set`` objects to include in the diagram.

        ## Example:

            v = IVenn(
                Set("A", [1, 2, 3]),
                Set("B", [3, 4, 5]),
            )
        """
        if len(sets) == 0:
            raise ValueError("You must provide at least one Set.")
        if len(sets) > 6:
            raise ValueError("This implementation supports up to 6 sets.")

        self.sets: dict[str, set[str]] = {}
        self.labels: dict[str, str] = {}
        self.descriptions: dict[str, str] = {}

        for index, input_set in enumerate(sets):
            letter = chr(ord("A") + index)
            input_set.name = letter

            elements: set[str] = set()
            for element in input_set.elements:
                if element is None:
                    continue

                value = str(element).strip()
                if value:
                    elements.add(value)

            self.sets[letter] = elements
            self.labels[letter] = str(input_set.label).strip() if input_set.label else letter
            self.descriptions[letter] = str(getattr(input_set, "desc", "")).strip()

        self.union_states: list[list[set[str]]] = [[]]
        self.current_index: int = 0
        self.union_mode: str = "list"
        self._stopped: bool = False

        self.theme: str = "Default"
        self.custom_theme: dict[str, str] | None = None
        self.font_scale: float = 1.0
        self.opacity_scale: float = 1.0
        self.show_percentages: bool = False

        self._viewer = None

    # ------------------------- Constructors -----------------------------

    @classmethod
    def from_excel(cls, path: str, has_labels: bool = True) -> "IVenn":
        """Create an ``IVenn`` object from an Excel file.
        
        :param path: Path to the Excel file.
        :param has_labels: Whether the first row contains set labels. If ``True``,
            column headers are used as set labels. If ``False``, internal labels
            ``A``, ``B``, ``C`` ... are used instead. (default is True)
            
        ## Example:
        
            v = IVenn.from_excel("sets.xlsx")
            v = IVenn.from_excel("sets.xlsx", has_labels=False)
        """
        if not os.path.exists(path):
            raise FileNotFoundError(f"Excel file not found: {path}")

        if has_labels:
            dataframe = pd.read_excel(path, dtype=str)
        else:
            dataframe = pd.read_excel(path, dtype=str, header=None)

        def normalise(value: object) -> str | None:
            if pd.isna(value):
                return None
            text = str(value).strip()
            if not text:
                return None
            if text.endswith(".0"):
                text = text[:-2]
            return text

        excel_sets: list[Set] = []

        for index, column in enumerate(dataframe.columns):
            elements: set[str] = set()
            for value in dataframe[column]:
                element = normalise(value)
                if element:
                    elements.add(element)

            label = str(column).strip() if has_labels else chr(ord("A") + index)
            excel_sets.append(Set(label, elements))

        return cls(*excel_sets)

    # ------------------------- Public themes -----------------------------

    @staticmethod
    def available_themes() -> list[str]:
        """Return the names of the available themes.

        :return: A sorted list of available theme names.

        ## Example:

            IVenn.available_themes()
        """
        return theme_names()

    def set_theme(self, *args) -> None:
        """Apply a built-in theme or define a custom six-colour theme.

        Pass either the name of a built-in theme, or six hex colours for sets
        ``A`` to ``F``.

        :param args: Either one theme name or six hex colour strings.

        ## Example:

            v.set_theme("Default")
            v.set_theme("#111111", "#222222", "#333333", "#444444", "#555555", "#666666")
        """
        if len(args) == 1 and isinstance(args[0], str):
            name = args[0].strip()

            if name.lower() == "custom":
                self.theme = "Custom" if self.custom_theme else "Default"
                self._render_if_viewer()
                return

            if name not in SET_COLOUR_THEMES:
                raise ValueError(f"Unknown theme: {name}")

            self.theme = name
            self._render_if_viewer()
            return

        if len(args) == 6 and all(isinstance(colour, str) for colour in args):
            colours = [colour.strip() for colour in args]
            hex_pattern = re.compile(r"^#([0-9a-fA-F]{6}|[0-9a-fA-F]{8})$")

            if not all(hex_pattern.match(colour) for colour in colours):
                raise ValueError("Colours must be hex strings like #RRGGBB (or #RRGGBBAA).")

            self.custom_theme = {
                "A": colours[0],
                "B": colours[1],
                "C": colours[2],
                "D": colours[3],
                "E": colours[4],
                "F": colours[5],
                "_opacity": SET_COLOUR_THEMES["Default"].get("_opacity", 0.4),
            }
            validate_theme(self.custom_theme)
            self.theme = "Custom"
            self._render_if_viewer()
            return

        raise TypeError("set_theme expects set_theme('ThemeName') or 6 hex colours.")


    # ------------------------- Public union configuration -----------------------------

    def set_unions(self, unions: str | None) -> None:
        """Define the union states used by the current diagram view.

        The input may be either list-style unions such as ``"ab; cde,f"`` or
        tree-style unions such as ``"((A,B),C)"``. Passing an empty value clears
        all unions and returns to the base view.

        :param unions: Union definition string, or ``None`` to clear unions.

        ## Example:

            v.set_unions("ab; cde,f")
            v.set_unions("((A,B),C)")
        """
        self._stopped = False
        text = "" if unions is None else str(unions).strip()
        if not text:
            self.union_mode = "list"
            self.union_states = [[]]
            self.current_index = 0
            self._render_if_viewer()
            return

        if "(" in text and ")" in text:
            self._set_unions_tree(text)
        else:
            self._set_unions_list(text)

        self.current_index = 0
        self._render_if_viewer()
        

    # ------------------------- Public union navigation methods -----------------------------
    
    def union_views(self) -> list[str]:
        """Return the stored union views.

        The base non-unioned view is returned as ``"base"``. Other views are
        returned as comma-separated union names such as ``"AB"`` or ``"CD,EF"``.

        :return: A list of available union view names.

        ## Example:

            v.union_views()
        """
        views: list[str] = []

        for union_view in self.union_states:
            if not union_view:
                views.append("base")
                continue

            names = ["".join(sorted(group)) for group in union_view]
            names.sort()
            views.append(",".join(names))

        return views


    def current_union_view(self) -> str:
        """Return the name of the currently active union view.

        :return: The current union view name.

        ## Example:

            v.current_union_view()
        """
        return self.union_views()[self.current_index]


    def goto_union_view(self, unions: str | None) -> None:
        """Move to a stored union view by name.

        Pass ``None``, ``""``, or ``"base"`` to return to the non-unioned view.
        Otherwise, pass a single union-view name such as ``"AB"`` or ``"CD,EF"``.

        :param unions: Name of the union view to activate.

        ## Example:

            v.goto_union_view("AB")
            v.goto_union_view("CD,EF")
            v.goto_union_view("base")
        """
        target = self._normalise_union_view_name(unions)

        for index, name in enumerate(self.union_views()):
            if name == target:
                self._stopped = False
                self.current_index = index
                self._render_if_viewer()
                return

        raise ValueError(f"Unknown union view: {unions!r}")


    def has_union_view(self, unions: str | None) -> bool:
        """Return whether a named union view exists.

        :param unions: Name of the union view to check.
        :return: ``True`` if the view exists, else ``False``.

        ## Example:

            v.has_union_view("AB")
            v.has_union_view("CD,EF")
        """
        target = self._normalise_union_view_name(unions)
        return target in self.union_views()


    def reset_union_view(self) -> None:
        """Return to the base non-unioned view (without deleting stored unions).

        ## Example:

            v.reset_union_view()
        """
        self.goto_union_view("base")


    # ------------------------ Public analysis methods -----------------------------

    def intersections(self, min_size: int = 1, min_degree: int = 1, max_degree: int | None = None, order_by: str = "size", top: int | None = None, include_empty: bool = False, include_elements: bool = False) -> list[dict[str, object]]:
        """Return filtered intersections for the current view.

        Each record contains a user-facing intersection label, its size, its
        degree, and, if requested, its sorted elements.

        :param min_size: Minimum intersection size to include.
        :param min_degree: Minimum intersection degree to include.
        :param max_degree: Maximum intersection degree to include.
        :param order_by: Sort order for the returned records, either by 'size' or 'degree'.
        :param top: Maximum number of records to return.
        :param include_empty: Include intersections of size 0.
        :param include_elements: Include the elements in each returned intersection.
        :return: A list of user-facing intersection records.

        ## Example:

            v.intersections()
            v.intersections(min_degree=2, include_elements=True)
        """
        records = self._intersections(
            min_size=min_size,
            min_degree=min_degree,
            max_degree=max_degree,
            order_by=order_by,
            top=top,
            include_empty=include_empty,
            include_elements=include_elements,
        )

        public_records: list[dict[str, object]] = []
        for record in records:
            public_record = {
                "intersection": record["label"],
                "size": record["size"],
                "degree": record["degree"],
            }
            if include_elements:
                public_record["elements"] = record["elements"]
            public_records.append(public_record)

        return public_records

    def top_intersections(self, top: int = 10, min_size: int = 1, min_degree: int = 1, max_degree: int | None = None, order_by: str = "size", include_empty: bool = False, include_elements: bool = False) -> list[dict[str, object]]:
        """Return the top intersection records from the current view.

        :param top: Maximum number of records to return.
        :param min_size: Minimum intersection size to include.
        :param min_degree: Minimum intersection degree to include.
        :param max_degree: Maximum intersection degree to include.
        :param order_by: Sort order for the returned records, either by 'size' or 'degree'.
        :param include_empty: Include intersections of size 0.
        :param include_elements: Include the elements in each returned intersection.
        :return: A list of the top intersection records.

        ## Example:

            v.top_intersections(top=5)
        """
        return self.intersections(
            min_size=min_size,
            min_degree=min_degree,
            max_degree=max_degree,
            order_by=order_by,
            top=top,
            include_empty=include_empty,
            include_elements=include_elements,
        )

    def get_intersection(self, intersection: str) -> set[str]:
        """Return the elements of one intersection.

        The input may be either an internal key such as ``"AB"`` or a user-facing
        intersection label such as ``"Patients ∩ Controls"``.

        :param intersection: Intersection to look up.
        :return: The elements in that intersection, or an empty set if not found.

        ## Example:

            v.get_intersection("AB")
            v.get_intersection("Patients ∩ Controls")
        """
        key = self._normalise_intersection_lookup(intersection)
        return self._region_elements().get(key, set())

    def empty_intersections(self) -> list[str]:
        """Return the labels of all empty intersections in the current view.

        :return: A sorted list of user-facing intersection labels with size 0.

        ## Example:

            v.empty_intersections()
        """
        empty_labels: list[str] = []
        sorted_regions = sorted(
            self._region_sizes().items(),
            key=lambda item: len(self._split_region_key(item[0])),
        )

        for key, size in sorted_regions:
            if size == 0:
                empty_labels.append(self._format_region_label(key))

        return empty_labels

    def non_empty_intersections(self) -> list[str]:
        """Return the labels of all non-empty intersections in the current view.

        :return: A sorted list of user-facing intersection labels with size > 0.

        ## Example:

            v.non_empty_intersections()
        """
        non_empty_labels: list[str] = []
        sorted_regions = sorted(
            self._region_sizes().items(),
            key=lambda item: len(self._split_region_key(item[0])),
        )

        for key, size in sorted_regions:
            if size > 0:
                non_empty_labels.append(self._format_region_label(key))

        return non_empty_labels


    # ------------------------ Public export / rendering methods -----------------------------

    def export_svg(self, path: str, base_dir: str | None = None) -> str:
        """Export the current diagram view as an SVG file.

        :param path: Output path for the SVG file.
        :param base_dir: Optional directory containing diagram templates.
        :return: The output path.

        ## Example:

            v.export_svg("diagram.svg")
        """
        return self._render_svg(path, base_dir=base_dir)

    def export_png(self, path: str, base_dir: str | None = None, scale: float = 2.0) -> str:
        """Export the current diagram view as a PNG file.

        :param path: Output path for the PNG file.
        :param base_dir: Optional directory containing diagram templates.
        :param scale: Requested render scale.
        :return: The output path.

        ## Example:

            v.export_png("diagram.png")
        """
        try:
            from ivenn.gui.launcher import export_png
        except ImportError as exc:
            raise ImportError(
                "PNG export requires the GUI launcher to be available in the installed package."
            ) from exc
        return export_png(self, path, base_dir=base_dir, scale=scale)

    def export_sets(self, path: str) -> str:
        """Export the original input sets to an Excel file.

        :param path: Output path for the Excel file.
        :return: The output path.

        ## Example:

            v.export_sets("sets.xlsx")
        """
        return self._export_sets(path)

    def export_intersections(self, path: str, min_size: int = 1, min_degree: int = 1, max_degree: int | None = None, order_by: str = "degree_then_key", top: int | None = None) -> str:
        """Export intersection elements to an Excel workbook.

        Each exported column represents one region, and the rows below it contain
        the elements in that region.

        :param path: Output path for the Excel file.
        :param min_size: Minimum region size to include.
        :param min_degree: Minimum intersection degree to include.
        :param max_degree: Maximum intersection degree to include.
        :param order_by: Sort order for exported regions.
        :param top: Maximum number of regions to export.
        :return: The output path.

        ## Example:

            v.export_intersections("intersections.xlsx", min_degree=2)
        """
        return self._export_intersections(
            path,
            min_size=min_size,
            min_degree=min_degree,
            max_degree=max_degree,
            order_by=order_by,
            top=top,
        )

    def draw(self) -> None:
        """Open the GUI viewer for the current ``IVenn`` instance.

        ## Example:

            v.draw()
        """
        try:
            from ivenn.gui.launcher import render_controller
        except ImportError as exc:
            raise ImportError(
                "GUI support is not available in the installed package."
            ) from exc
        render_controller(self, start=True)


    # ------------------------ Internal viewer helpers -----------------------------

    def _render_if_viewer(self) -> None:
        """Re-render the diagram if a viewer window is already open."""
        if self._viewer is None:
            return
        from ivenn.gui.launcher import render_controller
        render_controller(self, start=False)
        
    def _font(self, scale: float) -> None:
        self.font_scale = max(0.4, min(3.0, float(scale)))
        self._render_if_viewer()

    def _opacity(self, scale: float) -> None:
        self.opacity_scale = max(0.5, min(2.0, float(scale)))
        self._render_if_viewer()

    def _set_show_percentages(self, enabled: bool) -> None:
        """Toggle percentage display for region labels in the viewer."""
        self.show_percentages = bool(enabled)
        self._render_if_viewer()

    def _stop(self) -> None:
        """Reset navigation to the base non-unioned view."""
        self._stopped = True
        self.union_states = [[]]
        self.current_index = 0
        self._render_if_viewer()

    def _has_unions(self) -> bool:
        """Return whether the current controller has additional union states."""
        return (not self._stopped) and len(self.union_states) > 1

    def _next(self) -> None:
        """Advance to the next union state if one exists."""
        if self._stopped:
            return
        if self.current_index < len(self.union_states) - 1:
            self.current_index += 1
            self._render_if_viewer()

    def _prev(self) -> None:
        """Move to the previous union state if one exists."""
        if self._stopped:
            return
        if self.current_index > 0:
            self.current_index -= 1
            self._render_if_viewer()


    # ------------------------ Internal union parsing -----------------------------

    def _set_unions_list(self, union_string: str) -> None:
        """Parse and store a list-style union definition string."""
        self.union_mode = "list"
        self.union_states = [[]]

        for state_text in union_string.split(";"):
            state_text = state_text.strip()
            if not state_text:
                continue

            used_letters: set[str] = set()
            unions_for_state: list[set[str]] = []

            for union_text in state_text.split(","):
                letters = {character.upper() for character in union_text if character.isalpha()}
                if not letters:
                    continue

                unknown_letters = letters - set(self.sets.keys())
                if unknown_letters:
                    raise ValueError(f"Unknown set(s) in unions: {sorted(unknown_letters)}")

                if used_letters & letters:
                    raise ValueError("Overlapping unions are not allowed in the same state.")

                used_letters |= letters
                unions_for_state.append(set(letters))

            self.union_states.append(unions_for_state)

        if not self.union_states:
            self.union_states = [[]]

    def _tokenise_tree(self, text: str) -> list[str]:
        """Tokenise a tree-style union string into symbols and letters."""
        tokens: list[str] = []

        for character in text:
            if character.isspace():
                continue
            if character in "(),":
                tokens.append(character)
            elif character.isalpha():
                tokens.append(character.upper())
            else:
                raise ValueError(f"Invalid character in tree union: {character!r}")

        return tokens

    def _set_unions_tree(self, tree_string: str) -> None:
        """Parse and store a tree-style union definition string."""
        self.union_mode = "tree"

        tokens = self._tokenise_tree(tree_string)
        if not tokens:
            raise ValueError("Tree union input is empty.")

        stack: list[object] = []
        used_letters: set[str] = set()
        groups_with_depth: list[tuple[int, set[str]]] = []

        current_depth = 0
        expect_item = True

        for token in tokens:
            if token == "(":
                stack.append("(")
                current_depth += 1
                expect_item = True
            elif token == ",":
                if expect_item:
                    raise ValueError("Missing item before or after ',' in tree union.")
                expect_item = True
            elif token == ")":
                if expect_item:
                    raise ValueError("Empty group or trailing ',' in tree union.")

                items: list[set[str]] = []
                while stack and stack[-1] != "(":
                    item = stack.pop()
                    if not isinstance(item, set):
                        raise ValueError("Invalid tree union structure.")
                    items.append(item)

                if not stack:
                    raise ValueError("Unmatched ')' in tree union.")

                stack.pop()
                current_depth -= 1

                if len(items) < 2:
                    raise ValueError("Each tree union group must contain at least 2 items.")

                merged: set[str] = set()
                for item in items:
                    if merged & item:
                        raise ValueError("Overlapping groups are not allowed in tree union.")
                    merged |= item

                groups_with_depth.append((current_depth, merged))
                stack.append(merged)
                expect_item = False

            else:
                letter = token.upper()

                if letter not in self.sets:
                    raise ValueError(f"Unknown set in tree union: {letter}")
                if letter in used_letters:
                    raise ValueError(f"Set {letter} appears more than once in the tree union.")

                used_letters.add(letter)
                stack.append({letter})
                expect_item = False

        if expect_item:
            raise ValueError("Tree union ends unexpectedly.")
        if len(stack) != 1 or not isinstance(stack[0], set):
            raise ValueError("Invalid tree union structure.")

        root_group = stack[0]

        groups_by_depth: dict[int, list[set[str]]] = {}
        for depth, group in groups_with_depth:
            groups_by_depth.setdefault(depth, []).append(group)

        self.union_states = [[]]
        active_groups: list[set[str]] = []

        for depth in sorted(groups_by_depth.keys(), reverse=True):
            active_groups.extend(groups_by_depth[depth])

            visible_groups: list[set[str]] = []
            for group in active_groups:
                if not any(group < other for other in active_groups):
                    visible_groups.append(group)

            unions = [set(group) for group in visible_groups if len(group) > 1]
            unions.sort(key=lambda group: (len(group), "".join(sorted(group))))

            if unions != self.union_states[-1]:
                self.union_states.append(unions)

        final_union = [set(root_group)]
        if final_union != self.union_states[-1]:
            self.union_states.append(final_union)


    # ------------------------ Internal data helpers -----------------------------

    def _collapsed_sets(self) -> dict[str, set[str]]:
        """Return the sets after applying the current union state."""
        grouped_sets: dict[str, set[str]] = {}
        used_letters: set[str] = set()

        current_unions = self.union_states[self.current_index] if self.union_states else []
        for union in current_unions:
            key = "".join(sorted(union))
            elements: set[str] = set()
            for name in union:
                elements |= self.sets[name]
            grouped_sets[key] = elements
            used_letters |= union

        for name, elements in self.sets.items():
            if name not in used_letters:
                grouped_sets[name] = set(elements)

        return grouped_sets

    def _region_sizes(self) -> dict[str, int]:
        """Return the exclusive size of every region in the current view."""
        grouped_sets = self._collapsed_sets()
        names = sorted(grouped_sets.keys())
        region_sizes: dict[str, int] = {}

        for degree in range(1, len(names) + 1):
            for combination_names in combinations(names, degree):
                region_elements = grouped_sets[combination_names[0]].copy()
                for name in combination_names[1:]:
                    region_elements &= grouped_sets[name]
                for other_name in set(names) - set(combination_names):
                    region_elements -= grouped_sets[other_name]
                key = self._make_region_key(combination_names)
                region_sizes[key] = len(region_elements)

        return region_sizes

    def _region_elements(self) -> dict[str, set[str]]:
        """Return the exclusive elements of every region in the current view."""
        grouped_sets = self._collapsed_sets()
        names = sorted(grouped_sets.keys())
        region_elements_by_key: dict[str, set[str]] = {}

        for degree in range(1, len(names) + 1):
            for combination_names in combinations(names, degree):
                region_elements = grouped_sets[combination_names[0]].copy()
                for name in combination_names[1:]:
                    region_elements &= grouped_sets[name]
                for other_name in set(names) - set(combination_names):
                    region_elements -= grouped_sets[other_name]
                key = self._make_region_key(combination_names)
                region_elements_by_key[key] = region_elements

        return region_elements_by_key

    def _intersections(self, min_size: int = 1, min_degree: int = 1, max_degree: int | None = None, order_by: str = "size", top: int | None = None, include_empty: bool = False, include_elements: bool = False) -> list[dict[str, object]]:
        """Return filtered internal intersection records for the current view."""
        region_sizes = self._region_sizes()
        region_elements_by_key = self._region_elements()

        records: list[dict[str, object]] = []

        for key, size in region_sizes.items():
            degree = len(self._split_region_key(key))

            if not include_empty and size == 0:
                continue
            if size < min_size:
                continue
            if degree < min_degree:
                continue
            if max_degree is not None and degree > max_degree:
                continue

            record: dict[str, object] = {
                "key": key,
                "label": self._format_region_label(key),
                "size": size,
                "degree": degree,
            }

            if include_elements:
                record["elements"] = tuple(
                    sorted((str(element) for element in region_elements_by_key.get(key, set())), key=str)
                )

            records.append(record)

        self._sort_region_records(records, order_by=order_by)

        if top is not None:
            records = records[:top]

        return records

    def _label_to_letter_map(self) -> dict[str, str]:
        """Return a reverse mapping from user-facing set labels to internal letters."""
        mapping: dict[str, str] = {}
        duplicate_labels: set[str] = set()

        for letter, label in self.labels.items():
            if label in mapping and mapping[label] != letter:
                duplicate_labels.add(label)
            mapping[label] = letter

        if duplicate_labels:
            duplicate_text = ", ".join(sorted(duplicate_labels))
            raise ValueError(
                f"Set labels must be unique for label-based intersection lookup: {duplicate_text}"
            )

        return mapping

    def _normalise_intersection_lookup(self, intersection: str) -> str:
        """Return an internal region key from an internal key, SVG id, or display label."""
        text = str(intersection).strip()
        text = text.replace("(", "").replace(")", "")
        if not text:
            raise ValueError("Intersection cannot be empty.")

        region_map = self._region_elements()

        if "|" in text:
            groups = ["".join(sorted(part.upper())) for part in text.split("|") if part.strip()]
            key = "|".join(sorted(groups))
            if key in region_map:
                return key

        flat = "".join(character for character in text if character.isalpha()).upper()
        flat_map = self._template_region_lookup_map()
        if flat in flat_map:
            return flat_map[flat]

        label_map = self._label_to_letter_map()
        if "∩" in text or "∪" in text:
            groups: list[str] = []
            for group_text in text.split("∩"):
                letters: list[str] = []
                for label_text in group_text.split("∪"):
                    label = label_text.strip()
                    if not label:
                        continue
                    if label not in label_map:
                        raise ValueError(f"Unknown set label in intersection: {label}")
                    letters.append(label_map[label])
                if letters:
                    groups.append("".join(sorted(letters)))

            key = "|".join(sorted(groups))
            if key in region_map:
                return key
            raise ValueError(f"Unknown intersection: {intersection!r}")
        
        if text in label_map:
            key = label_map[text]
            if key in region_map:
                return key

        raise ValueError(f"Unknown intersection: {intersection!r}")
    
    @staticmethod
    def _normalise_union_view_name(unions: str | None) -> str:
        """Return a name for one union view."""
        if unions is None:
            return "base"

        text = str(unions).strip()
        if not text or text.lower() == "base":
            return "base"

        parts: list[str] = []

        for part in text.split(","):
            letters = sorted({character.upper() for character in part if character.isalpha()})
            if letters:
                parts.append("".join(letters))

        if not parts:
            return "base"

        parts.sort()
        return ",".join(parts)

    @staticmethod
    def _normalise_key(parts: tuple[str, ...]) -> str:
        """Return a sorted region key built from one or more set names."""
        return "".join(sorted("".join(parts)))

    def _format_region_label(self, raw_key: str) -> str:
        """Convert an internal region key into a display label."""
        groups = self._split_region_key(raw_key)
        if not groups:
            return ""

        formatted_groups: list[str] = []
        multiple_groups = len(groups) > 1

        for group in groups:
            labels = [self.labels.get(letter, letter) for letter in group]

            if len(labels) == 1:
                formatted_groups.append(labels[0])
            else:
                union_text = " ∪ ".join(labels)
                if multiple_groups:
                    union_text = f"({union_text})"
                formatted_groups.append(union_text)

        return " ∩ ".join(formatted_groups)
    
    @staticmethod
    def _make_region_key(parts: tuple[str, ...]) -> str:
        """Return a internal region key preserving union-group structure."""
        groups = ["".join(sorted(part)) for part in parts]
        groups.sort()
        return "|".join(groups)

    @staticmethod
    def _split_region_key(key: str) -> tuple[str, ...]:
        """Split an internal region key into its visible-group parts."""
        if not key:
            return ()
        return tuple(part for part in key.split("|") if part)

    @staticmethod
    def _template_region_id(parts: tuple[str, ...]) -> str:
        """Return the SVG template id for a combination of visible groups."""
        return "".join(sorted("".join(parts))).lower()
    
    def _template_region_lookup_map(self) -> dict[str, str]:
        """Map SVG region ids to structured internal region keys."""
        names = sorted(self._collapsed_sets().keys())
        mapping: dict[str, str] = {}

        for degree in range(1, len(names) + 1):
            for combination_names in combinations(names, degree):
                flat = self._template_region_id(combination_names).upper()
                structured = self._make_region_key(combination_names)
                mapping[flat] = structured

        return mapping

    @staticmethod
    def _sort_region_records(records: list[dict[str, object]], order_by: str) -> None:
        """Sort region records in place using one of the supported sort modes."""
        if order_by == "degree":
            records.sort(key=lambda item: (item["degree"], -item["size"], item["key"]))
        elif order_by == "key":
            records.sort(key=lambda item: item["key"])
        elif order_by == "size":
            records.sort(key=lambda item: item["size"])
        else:
            records.sort(key=lambda item: item["size"])


    # ------------------------ Internal rendering / export helpers -----------------------------

    @contextmanager
    def _diagram_template_path(self, base_dir: str | None = None):
        """Give the SVG template pathe."""
        number_of_sets = len(self.sets)
        unions = self.union_states[self.current_index] if self.union_states else []

        if not unions:
            filename = f"{number_of_sets}waydiagram.svg"
        else:
            union_names = ["".join(sorted(union)).lower() for union in unions]
            union_names.sort()
            suffix = "_".join(union_names)
            filename = f"{number_of_sets}waydiagram_{suffix}.svg"

        if base_dir:
            yield os.path.join(base_dir, str(number_of_sets), filename)
            return

        resource = files("ivenn").joinpath("diagrams", str(number_of_sets), filename)
        with as_file(resource) as template_path:
            yield str(template_path)

    def _build_replacements(self) -> dict[str, object]:
        """Build SVG text replacements for labels, totals, and region counts."""
        replacements: dict[str, object] = {}

        for name, elements in self.sets.items():
            replacements[f"label{name}"] = self.labels.get(name, name)
            replacements[f"total{name}"] = f"({len(elements)})"

        region_sizes = self._region_sizes()
        flat_counts: dict[str, int] = {}

        for key, count in region_sizes.items():
            parts = self._split_region_key(key)
            region_id = self._template_region_id(parts)
            flat_counts[region_id] = count

        if self.show_percentages:
            total = sum(flat_counts.values())
            for region_id in self._all_region_ids():
                count = flat_counts.get(region_id, 0)
                percentage = round((count / total) * 100) if total > 0 else 0
                replacements[region_id] = f"{percentage}%"
        else:
            for region_id in self._all_region_ids():
                replacements[region_id] = flat_counts.get(region_id, 0)

        return replacements

    def _all_region_ids(self) -> list[str]:
        """Return all region ids expected by the current diagram template."""
        names = sorted(self._collapsed_sets().keys())
        region_ids: list[str] = []

        for degree in range(1, len(names) + 1):
            for combination_names in combinations(names, degree):
                region_ids.append(self._normalise_key(combination_names).lower())

        return region_ids

    def _apply_replacements(self, svg_template_path: str, replacements: dict[str, object], out_path: str) -> None:
        """Apply labels, counts, and theme styles to an SVG template."""
        if not os.path.exists(svg_template_path):
            raise FileNotFoundError(f"Template not found: {svg_template_path}")

        tree = etree.parse(svg_template_path)
        root = tree.getroot()

        if self.theme == "Custom" and self.custom_theme:
            theme = self.custom_theme
            base_opacity = SET_COLOUR_THEMES["Default"].get("_opacity", 0.4)
        else:
            theme = get_theme(self.theme)
            base_opacity = theme.get("_opacity", 0.4)

        opacity = max(0.05, min(1.0, float(base_opacity) * self.opacity_scale))

        for path in root.xpath(".//svg:path", namespaces=SVG_NS):
            path_id = path.get("id", "")
            match = re.match(r"ellipse([A-F])$", path_id, re.IGNORECASE)
            if not match:
                continue

            set_name = match.group(1).upper()
            if set_name not in theme:
                continue

            style = path.get("style", "")
            style = self._set_fill_style(style, str(theme[set_name]), opacity)
            path.set("style", style)

        for text in root.xpath(".//svg:text", namespaces=SVG_NS):
            node_id = text.get("id")
            if not node_id:
                continue

            for child in list(text):
                text.remove(child)

            style = text.get("style", "")
            style = self._set_text_style(style, self.font_scale)

            if node_id.startswith("label") or node_id.startswith("total"):
                set_name = node_id[-1].upper()
                if set_name in theme:
                    style = re.sub(r"fill:[^;]+;?", "", style)
                    style += ";text-anchor:middle"
                    style += f";fill:{theme[set_name]}"

            text.set("style", style)
            text.text = str(replacements.get(node_id, ""))

        tree.write(out_path)

    @staticmethod
    def _set_fill_style(style: str, colour: str, opacity: float) -> str:
        """Return an SVG style string with updated fill colour and opacity."""
        parts = [
            part
            for part in style.split(";")
            if not part.strip().startswith(("fill:", "fill-opacity"))
        ]
        parts.append(f"fill:{colour}")
        parts.append(f"fill-opacity:{opacity}")
        return ";".join(parts)

    @staticmethod
    def _set_text_style(style: str, scale: float) -> str:
        """Return an SVG text style string with an updated font size."""
        base_size = 14 * scale
        parts = [part for part in style.split(";") if not part.strip().startswith("font-size")]
        parts.append(f"font-size:{base_size}px")
        return ";".join(parts)

    def _render_svg(self, out_path: str, base_dir: str | None = None) -> str:
        """Render the current diagram view to an SVG file."""
        replacements = self._build_replacements()
        with self._diagram_template_path(base_dir=base_dir) as template_path:
            self._apply_replacements(template_path, replacements, out_path)
        return out_path

    def _export_sets(self, path: str) -> str:
        """Write the original input sets to an Excel workbook."""
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        columns: dict[str, list[str]] = {}
        max_length = max((len(values) for values in self.sets.values()), default=0)

        for name, elements in self.sets.items():
            column = sorted((str(element) for element in elements), key=str)
            column += [""] * (max_length - len(column))
            columns[self.labels.get(name, name)] = column

        dataframe = pd.DataFrame(columns)
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            dataframe.to_excel(writer, sheet_name="Sets", index=False)

        return path

    def _export_intersections(self, path: str, min_size: int = 1, min_degree: int = 1, max_degree: int | None = None, order_by: str = "degree_then_key", top: int | None = None) -> str:
        """Write intersection elements to an Excel workbook."""
        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        records = self._intersections(
            min_size=min_size,
            min_degree=min_degree,
            max_degree=max_degree,
            order_by=order_by,
            top=top,
            include_elements=True,
        )

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Intersections"

        used_headers: set[str] = set()
        column_index = 1

        for record in records:
            header = record["label"]
            base_header = header
            suffix = 2
            while header in used_headers:
                header = f"{base_header} ({suffix})"
                suffix += 1
            used_headers.add(header)

            worksheet.cell(row=1, column=column_index, value=header)
            for row_index, element in enumerate(record["elements"], start=2):
                worksheet.cell(row=row_index, column=column_index, value=element)

            column_index += 1

        if column_index == 1:
            worksheet.cell(row=1, column=1, value="No intersections")

        workbook.save(path)
        return path